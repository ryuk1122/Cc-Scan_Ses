"""
CedulaScan Pro — FastAPI backend with afiliados DB + admin features
Real-time scanning of Colombian cedulas with zero-duplicate guarantees + autofill.

Correcciones v2.1:
  - Reemplazado @app.on_event (deprecado) por lifespan context manager.
  - Índice TTL de 24 h en colección deduplication (evita crecimiento infinito).
  - Índice único en deduplication.idempotency_key.
  - Separación clara de startup / shutdown.
"""
from contextlib import asynccontextmanager
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Header, Query, WebSocket, WebSocketDisconnect, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr, Field
from pymongo.errors import DuplicateKeyError
from pymongo import UpdateOne
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any
from pathlib import Path
import asyncio
import hashlib
import io
import logging
import os
import uuid
import csv as csv_lib
import jwt

from cedula_parser import clean_cedula, extract_cedula, parse_pdf417
from ocr_service import ocr_from_base64_async, ocr_dependency_status
from barcode_service import decode_identity_document_from_base64, barcode_dependency_status
from gemini_service import gemini_dependency_status

# ---------- Config & init ----------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")


def require_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Variable de entorno requerida no configurada: {name}")
    return value


def parse_csv_env(name: str, default: str = "") -> List[str]:
    return [item.strip() for item in os.environ.get(name, default).split(",") if item.strip()]


def parse_bool_env(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on", "si", "sí"}


MONGO_URL   = require_env("MONGO_URL")
DB_NAME     = require_env("DB_NAME")
JWT_SECRET  = require_env("JWT_SECRET_KEY")
JWT_ALGO    = os.environ.get("JWT_ALGORITHM", "HS256")
TOKEN_HOURS = int(os.environ.get("ACCESS_TOKEN_EXPIRE_HOURS", "12"))
CORS_ORIGINS = parse_csv_env("CORS_ORIGINS", "*")
BCRYPT_ROUNDS = int(os.environ.get("BCRYPT_ROUNDS", "10"))
ALLOW_WIPE_AFILIADOS = parse_bool_env("ALLOW_WIPE_AFILIADOS", False)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("cedulascan")

client = AsyncIOMotorClient(MONGO_URL)
db     = client[DB_NAME]
pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto", bcrypt__rounds=BCRYPT_ROUNDS)


# ---------- Pydantic models ----------
class UserRegister(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6, max_length=128)
    nombre: str = Field(min_length=1, max_length=80)


class UserLogin(BaseModel):
    email: EmailStr
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: Dict[str, Any]


class EventoCreate(BaseModel):
    nombre: str = Field(min_length=1, max_length=120)
    fecha: str
    lugar: Optional[str] = ""
    descripcion: Optional[str] = ""


class EventoOut(BaseModel):
    id: str
    nombre: str
    fecha: str
    lugar: str
    descripcion: str
    activo: bool
    created_at: str
    total_registros: int = 0


class EscanearRequest(BaseModel):
    cedula: str = Field(min_length=1, max_length=400)
    evento_id: str
    device_id: str
    timestamp: int
    nonce: str
    nombre: Optional[str] = ""
    sede: Optional[str] = ""
    municipio: Optional[str] = ""
    cargo: Optional[str] = ""
    raw_barcode: Optional[str] = ""
    primer_apellido: Optional[str] = ""
    segundo_apellido: Optional[str] = ""
    nombres: Optional[str] = ""
    genero: Optional[str] = ""
    fecha_nacimiento: Optional[str] = ""
    fecha_expiracion: Optional[str] = ""
    tipo_sangre: Optional[str] = ""


class AfiliadoUpdate(BaseModel):
    nombre: Optional[str] = None
    sede: Optional[str] = None
    municipio: Optional[str] = None
    zona: Optional[str] = None
    cargo: Optional[str] = None
    titulo: Optional[str] = None
    email: Optional[str] = None
    celular: Optional[str] = None
    fecha_nac: Optional[str] = None


class RegistroUpdate(BaseModel):
    nombre: Optional[str] = None
    sede: Optional[str] = None
    municipio: Optional[str] = None
    cargo: Optional[str] = None


# ---------- Helpers ----------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_password(pw: str) -> str:
    return pwd_ctx.hash(pw)


def verify_password(pw: str, h: str) -> bool:
    try:
        return pwd_ctx.verify(pw, h)
    except Exception:
        return False


def password_needs_update(h: str) -> bool:
    try:
        return pwd_ctx.needs_update(h)
    except Exception:
        return False


async def update_password_hash(user_id: str, password: str) -> None:
    try:
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"password_hash": hash_password(password)}},
        )
    except Exception as exc:
        logger.warning("No se pudo actualizar el hash de password: %s", exc)


def create_token(user_id: str, email: str, role: str) -> str:
    exp = datetime.now(timezone.utc) + timedelta(hours=TOKEN_HOURS)
    return jwt.encode(
        {"sub": user_id, "email": email, "role": role, "exp": exp},
        JWT_SECRET,
        algorithm=JWT_ALGO,
    )


def decode_token(token: str) -> Dict[str, Any]:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")


async def get_user_from_token(token: str) -> Dict[str, Any]:
    payload = decode_token(token)
    user_id = payload.get("sub")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no existe")
    return user


async def get_current_user(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Falta header Authorization")
    token = authorization.split(" ", 1)[1].strip()
    return await get_user_from_token(token)


async def require_admin(current=Depends(get_current_user)) -> Dict[str, Any]:
    if current.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Se requieren permisos de administrador")
    return current


async def require_admin_for_download(
    authorization: Optional[str] = Header(None),
    token_query: Optional[str] = Query(None, alias="_t"),
) -> Dict[str, Any]:
    if authorization and authorization.lower().startswith("bearer "):
        token = authorization.split(" ", 1)[1].strip()
    elif token_query:
        token = token_query.strip()
    else:
        raise HTTPException(status_code=401, detail="Falta token de autenticación")
    user = await get_user_from_token(token)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Se requieren permisos de administrador")
    return user


def compute_dedup_hash(cedula: str, evento_id: str, device_id: str, nonce: str) -> str:
    raw = f"{cedula}|{evento_id}|{device_id}|{nonce}"
    return hashlib.sha256(raw.encode()).hexdigest()


def clean_display_text(value: Any) -> str:
    text = str(value or "")
    text = "".join(ch if ch.isalnum() or ch.isspace() or ch in ".'-" else " " for ch in text)
    return " ".join(text.split()).strip()


TRUSTED_BARCODE_FORMATS = {
    "pdf417_binario_posicional",
    "pdf417_binario_compacto",
    "mrz_td1",
    "cedula_amarilla_guiones",
    "clasico_guiones",
    "pipe_separado",
    "numero_puro",
}


def trusted_barcode_cedula(parsed: Dict[str, Any]) -> Optional[str]:
    if not isinstance(parsed, dict):
        return None
    cedula = clean_cedula(parsed.get("cedula"))
    if not cedula:
        return None
    formato = str(parsed.get("formato_detectado") or "")
    if parsed.get("mrz_valido") or parsed.get("raw_mrz") or formato in TRUSTED_BARCODE_FORMATS:
        return cedula
    return None


# ---------- WebSocket manager ----------
class ConnectionManager:
    def __init__(self) -> None:
        self.rooms: Dict[str, List[WebSocket]] = {}
        self.devices: Dict[int, Dict[str, str]] = {}
        self.lock = asyncio.Lock()

    async def connect(self, ws: WebSocket, evento_id: str, device_id: str) -> None:
        await ws.accept()
        async with self.lock:
            self.rooms.setdefault(evento_id, []).append(ws)
            self.devices[id(ws)] = {"evento_id": evento_id, "device_id": device_id}
        await self.broadcast(evento_id, {
            "type": "device:conectado",
            "device_id": device_id,
            "activos": self.active_count(evento_id),
            "ts": int(datetime.now(timezone.utc).timestamp() * 1000),
        })

    async def disconnect(self, ws: WebSocket) -> None:
        meta = self.devices.pop(id(ws), None)
        if not meta:
            return
        evento_id = meta["evento_id"]
        async with self.lock:
            if evento_id in self.rooms:
                self.rooms[evento_id] = [w for w in self.rooms[evento_id] if w is not ws]
                if not self.rooms[evento_id]:
                    self.rooms.pop(evento_id, None)
        try:
            await self.broadcast(evento_id, {
                "type": "device:desconectado",
                "device_id": meta["device_id"],
                "activos": self.active_count(evento_id),
                "ts": int(datetime.now(timezone.utc).timestamp() * 1000),
            })
        except Exception:
            pass

    def active_count(self, evento_id: str) -> int:
        return len(self.rooms.get(evento_id, []))

    async def broadcast(self, evento_id: str, message: Dict[str, Any]) -> None:
        dead: List[WebSocket] = []
        for ws in list(self.rooms.get(evento_id, [])):
            try:
                await ws.send_json(message)
            except Exception:
                dead.append(ws)
        for ws in dead:
            await self.disconnect(ws)


manager = ConnectionManager()


# ---------- Afiliados seed ----------
async def _seed_afiliados_if_empty():
    count = await db.afiliados.count_documents({})
    if count > 0:
        logger.info(f"Afiliados ya cargados: {count}")
        return
    seed_file = ROOT_DIR / "seed" / "afiliados.csv"
    if not seed_file.exists():
        logger.info("No hay archivo seed de afiliados")
        return
    try:
        items: List[Dict[str, Any]] = []
        with open(seed_file, "r", encoding="utf-8") as f:
            reader = csv_lib.DictReader(f, delimiter=";")
            for row in reader:
                ced = clean_cedula(row.get("cedula"))
                if not ced:
                    continue
                items.append({
                    "cedula": ced,
                    "nombre":    (row.get("nombre")    or "").strip(),
                    "sede":      (row.get("sede")      or "").strip(),
                    "municipio": (row.get("municipio") or "").strip(),
                    "zona":      (row.get("zona")      or "").strip(),
                    "cargo":     (row.get("cargo")     or "").strip(),
                    "titulo":    (row.get("titulo")    or "").strip(),
                    "email":     (row.get("email")     or "").strip(),
                    "celular":   (row.get("celular")   or "").strip(),
                    "fecha_nac": (row.get("fecha_nac") or "").strip(),
                    "created_at": now_iso(),
                })
        BATCH = 1000
        total_in = 0
        for i in range(0, len(items), BATCH):
            chunk = items[i:i + BATCH]
            ops = [UpdateOne({"cedula": d["cedula"]}, {"$set": d}, upsert=True) for d in chunk]
            res = await db.afiliados.bulk_write(ops, ordered=False)
            total_in += (res.upserted_count or 0) + (res.modified_count or 0)
        logger.info(f"Afiliados sembrados: {total_in} de {len(items)}")
    except Exception as e:
        logger.exception(f"Error seeding afiliados: {e}")


async def _seed_admin_if_needed():
    """First-time admin from env vars ADMIN_EMAIL / ADMIN_PASSWORD."""
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@cedulascan.local")
    admin_pwd   = os.environ.get("ADMIN_PASSWORD", "Admin123!")
    existing    = await db.users.find_one({"email": admin_email})
    if existing:
        if existing.get("role") != "admin":
            await db.users.update_one({"email": admin_email}, {"$set": {"role": "admin"}})
            logger.info(f"Usuario {admin_email} promovido a admin")
        return
    doc = {
        "id":            str(uuid.uuid4()),
        "email":         admin_email,
        "nombre":        "Administrador",
        "password_hash": hash_password(admin_pwd),
        "role":          "admin",
        "created_at":    now_iso(),
    }
    await db.users.insert_one(doc)
    logger.info(f"Admin seed creado: {admin_email}")


# ---------- Setup / teardown ----------
async def _setup() -> None:
    """Crea índices y ejecuta seeds al arrancar."""
    # Usuarios
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id",    unique=True)
    # Eventos
    await db.eventos.create_index("id", unique=True)
    # Registros
    await db.registros.create_index("id", unique=True)
    await db.registros.create_index([("evento_id", 1), ("cedula", 1)], unique=True)
    await db.registros.create_index("dedup_hash", unique=True)
    await db.registros.create_index([("evento_id", 1), ("created_at", -1)])
    # Event store
    await db.event_store.create_index([("aggregate_id", 1), ("created_at", -1)])
    await db.event_store.create_index("dedup_hash")
    # Afiliados
    await db.afiliados.create_index("cedula", unique=True)
    await db.afiliados.create_index([("nombre", "text")])
    # Correccion: deduplicacion con TTL de 24 h e indice unico
    await db.deduplication.create_index("idempotency_key", unique=True)
    await db.deduplication.create_index(
        "created_at",
        expireAfterSeconds=86400,   # 24 horas: MongoDB limpia automaticamente
    )
    logger.info("Indices creados correctamente.")
    await _seed_admin_if_needed()
    asyncio.create_task(_seed_afiliados_if_empty())


# Correccion: lifespan reemplaza los deprecados @app.on_event
@asynccontextmanager
async def lifespan(app: FastAPI):
    await _setup()
    yield
    client.close()
    logger.info("Conexion MongoDB cerrada.")


# ---------- App ----------
app = FastAPI(title="CedulaScan Pro API", version="2.1.0", lifespan=lifespan)
api = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=CORS_ORIGINS != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------- Auth ----------
@api.post("/auth/register", response_model=TokenResponse, status_code=201)
async def register(body: UserRegister):
    existing = await db.users.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    user_id = str(uuid.uuid4())
    doc = {
        "id":            user_id,
        "email":         body.email.lower(),
        "nombre":        body.nombre,
        "password_hash": hash_password(body.password),
        "role":          "operator",
        "created_at":    now_iso(),
    }
    try:
        await db.users.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    token = create_token(user_id, doc["email"], doc["role"])
    return TokenResponse(
        access_token=token,
        user={"id": user_id, "email": doc["email"], "nombre": doc["nombre"], "role": doc["role"]},
    )


@api.post("/auth/login", response_model=TokenResponse)
async def login(body: UserLogin):
    user = await db.users.find_one({"email": body.email.lower()})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    if password_needs_update(user["password_hash"]):
        asyncio.create_task(update_password_hash(user["id"], body.password))
    token = create_token(user["id"], user["email"], user.get("role", "operator"))
    return TokenResponse(
        access_token=token,
        user={
            "id":     user["id"],
            "email":  user["email"],
            "nombre": user["nombre"],
            "role":   user.get("role", "operator"),
        },
    )


@api.get("/auth/me")
async def me(current=Depends(get_current_user)):
    return current


# ---------- Afiliados ----------
@api.get("/afiliados/{cedula}")
async def get_afiliado(cedula: str, current=Depends(get_current_user)):
    ced = clean_cedula(cedula)
    if not ced:
        raise HTTPException(status_code=400, detail="Cédula inválida")
    doc = await db.afiliados.find_one({"cedula": ced}, {"_id": 0})
    if not doc:
        return {"encontrado": False, "cedula": ced}
    return {"encontrado": True, "afiliado": doc}


@api.get("/admin/afiliados")
async def list_afiliados(q: str = "", limit: int = 50, skip: int = 0, current=Depends(require_admin)):
    query: Dict[str, Any] = {}
    qs = q.strip()
    if qs:
        if qs.isdigit():
            query = {"cedula": {"$regex": f"^{qs}"}}
        else:
            query = {"nombre": {"$regex": qs, "$options": "i"}}
    total = await db.afiliados.count_documents(query)
    items = await db.afiliados.find(query, {"_id": 0}).sort("nombre", 1).skip(skip).limit(min(limit, 200)).to_list(200)
    return {"total": total, "items": items}


@api.get("/admin/afiliados/stats")
async def afiliados_stats(current=Depends(require_admin)):
    total = await db.afiliados.count_documents({})
    return {"total": total}


@api.get("/admin/metrics")
async def admin_metrics(evento_id: str = "", current=Depends(require_admin)):
    event_filter: Dict[str, Any] = {"evento_id": evento_id} if evento_id else {}
    store_filter: Dict[str, Any] = {"aggregate_id": evento_id} if evento_id else {}
    now = datetime.now(timezone.utc)
    today_start = datetime(now.year, now.month, now.day, tzinfo=timezone.utc)
    today_ms = int(today_start.timestamp() * 1000)
    last_24h_ms = int((now - timedelta(hours=24)).timestamp() * 1000)

    total_afiliados, total_eventos, eventos_activos, total_registros, registros_evento, hoy, duplicados = await asyncio.gather(
        db.afiliados.count_documents({}),
        db.eventos.count_documents({}),
        db.eventos.count_documents({"activo": True}),
        db.registros.count_documents({}),
        db.registros.count_documents(event_filter) if event_filter else db.registros.count_documents({}),
        db.registros.count_documents({**event_filter, "timestamp": {"$gte": today_ms}}),
        db.event_store.count_documents({**store_filter, "event_type": "DUPLICADO_RECHAZADO"}),
    )

    afiliados_evento = await db.registros.count_documents({**event_filter, "es_afiliado": True})
    no_afiliados_evento = max(registros_evento - afiliados_evento, 0)
    con_nombre = await db.registros.count_documents({**event_filter, "nombre": {"$ne": ""}})
    con_fecha = await db.registros.count_documents({**event_filter, "fecha_nacimiento": {"$ne": ""}})
    dispositivos_usados = len(await db.registros.distinct("device_id", event_filter))

    async def aggregate_list(pipeline: List[Dict[str, Any]], limit: int = 8) -> List[Dict[str, Any]]:
        rows = await db.registros.aggregate(pipeline).to_list(limit)
        return rows

    top_municipios = await aggregate_list([
        {"$match": {**event_filter, "municipio": {"$nin": ["", None]}}},
        {"$group": {"_id": "$municipio", "total": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 8},
        {"$project": {"_id": 0, "label": "$_id", "total": 1}},
    ])
    top_operadores = await aggregate_list([
        {"$match": {**event_filter, "operator_email": {"$nin": ["", None]}}},
        {"$group": {"_id": "$operator_email", "total": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 8},
        {"$project": {"_id": 0, "label": "$_id", "total": 1}},
    ])
    por_hora = await aggregate_list([
        {"$match": {**event_filter, "timestamp": {"$gte": last_24h_ms}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%H:00", "date": {"$toDate": "$timestamp"}, "timezone": "America/Bogota"}},
            "total": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
        {"$project": {"_id": 0, "label": "$_id", "total": 1}},
    ], 24)

    recientes = await db.registros.find(event_filter, {"_id": 0}).sort("created_at", -1).limit(8).to_list(8)
    evento = await db.eventos.find_one({"id": evento_id}, {"_id": 0}) if evento_id else None
    return {
        "evento": evento,
        "global": {
            "afiliados": total_afiliados,
            "eventos": total_eventos,
            "eventos_activos": eventos_activos,
            "registros": total_registros,
        },
        "evento_actual": {
            "registros": registros_evento,
            "hoy": hoy,
            "afiliados": afiliados_evento,
            "no_afiliados": no_afiliados_evento,
            "duplicados": duplicados,
            "dispositivos_usados": dispositivos_usados,
            "dispositivos_activos": manager.active_count(evento_id) if evento_id else 0,
            "con_nombre": con_nombre,
            "con_fecha_nacimiento": con_fecha,
            "calidad_nombre_pct": round((con_nombre / registros_evento) * 100, 1) if registros_evento else 0,
            "calidad_fecha_pct": round((con_fecha / registros_evento) * 100, 1) if registros_evento else 0,
        },
        "top_municipios": top_municipios,
        "top_operadores": top_operadores,
        "por_hora": por_hora,
        "recientes": recientes,
        "ia": gemini_dependency_status(),
    }


@api.get("/admin/ia/status")
async def admin_ia_status(current=Depends(require_admin)):
    return {"gemini": gemini_dependency_status()}


@api.post("/admin/afiliados/import")
async def import_afiliados(file: UploadFile = File(...), current=Depends(require_admin)):
    """Importa afiliados desde CSV (con separador ; o ,) o XLSX. Upserts por cédula."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    fname = (file.filename or "").lower()
    items: List[Dict[str, Any]] = []
    try:
        if fname.endswith(".csv") or fname.endswith(".txt"):
            for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                try:
                    text = content.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("No se pudo decodificar el CSV")
            sample = text.splitlines()[0] if text else ""
            delim  = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv_lib.DictReader(io.StringIO(text), delimiter=delim)
            for raw_row in reader:
                row = {k.lower().strip(): (v or "").strip() for k, v in raw_row.items() if k}
                _ingest_afiliado_row(row, items)
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl no instalado")
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = [str(c).lower().strip() if c is not None else "" for c in next(rows_iter, [])]
            for r in rows_iter:
                row = {header[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(r) if i < len(header)}
                _ingest_afiliado_row(row, items)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer: {e}")

    if not items:
        raise HTTPException(status_code=400, detail="No se encontraron filas válidas con cédula")

    inserted = updated = 0
    BATCH = 500
    for i in range(0, len(items), BATCH):
        chunk = items[i:i + BATCH]
        ops = [
            UpdateOne(
                {"cedula": d["cedula"]},
                {"$set": d, "$setOnInsert": {"created_at": now_iso()}},
                upsert=True,
            )
            for d in chunk
        ]
        res = await db.afiliados.bulk_write(ops, ordered=False)
        inserted += res.upserted_count or 0
        updated  += res.modified_count or 0
    return {"ok": True, "total": len(items), "insertados": inserted, "actualizados": updated}


def _ingest_afiliado_row(row: Dict[str, str], items: List[Dict[str, Any]]) -> None:
    def pick(*aliases: str) -> str:
        for a in aliases:
            v = row.get(a.lower())
            if v:
                return v.strip()
        return ""

    nombres    = pick("nombres", "nombre", "nombre_completo", "first_name")
    apellidos  = pick("apellidos", "apellido", "last_name")
    nombre_full = pick("nombre_completo")
    if nombres or apellidos:
        nombre = f"{nombres} {apellidos}".replace(" - ", " ").replace(" -", "").strip()
    else:
        nombre = nombre_full

    ced_raw = pick("cedula", "cédula", "documento", "dni", "identificacion", "id")
    ced = clean_cedula(ced_raw)
    if not ced:
        return
    items.append({
        "cedula":    ced,
        "nombre":    nombre,
        "sede":      pick("sede", "institucion", "institución", "institucion_educativa"),
        "municipio": pick("municipio", "ente", "ciudad", "ciudad_trabajo"),
        "zona":      pick("zona", "mun. cedula", "mun. cédula", "mun_cedula"),
        "cargo":     pick("cargo", "rol"),
        "titulo":    pick("titulo", "título"),
        "email":     pick("email", "correo"),
        "celular":   pick("celular", "telefono", "teléfono", "movil", "móvil"),
        "fecha_nac": pick("fecha_nac", "fecha nac.", "fecha_nacimiento", "nacimiento"),
    })


@api.patch("/admin/afiliados/{cedula}")
async def update_afiliado(cedula: str, body: AfiliadoUpdate, current=Depends(require_admin)):
    ced = clean_cedula(cedula)
    if not ced:
        raise HTTPException(status_code=400, detail="Cédula inválida")
    data = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not data:
        raise HTTPException(status_code=400, detail="Sin cambios")
    res = await db.afiliados.update_one({"cedula": ced}, {"$set": data})
    if res.matched_count == 0:
        data["cedula"]     = ced
        data["created_at"] = now_iso()
        await db.afiliados.insert_one(data)
        return {"ok": True, "created": True}
    return {"ok": True, "updated": True}


@api.post("/admin/afiliados")
async def create_afiliado(body: AfiliadoUpdate, cedula: str = "", current=Depends(require_admin)):
    ced = clean_cedula(cedula)
    if not ced:
        raise HTTPException(status_code=400, detail="Cédula requerida")
    if await db.afiliados.find_one({"cedula": ced}):
        raise HTTPException(status_code=400, detail="Cédula ya existe")
    data = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    data["cedula"]     = ced
    data["created_at"] = now_iso()
    await db.afiliados.insert_one(data)
    return {"ok": True, "afiliado": {k: v for k, v in data.items() if k != "_id"}}


@api.delete("/admin/afiliados/{cedula}")
async def delete_afiliado(cedula: str, current=Depends(require_admin)):
    ced = clean_cedula(cedula)
    res = await db.afiliados.delete_one({"cedula": ced})
    return {"ok": True, "deleted": res.deleted_count}


@api.delete("/admin/afiliados")
async def wipe_afiliados(confirm: str = Query("", max_length=64), current=Depends(require_admin)):
    if not ALLOW_WIPE_AFILIADOS:
        raise HTTPException(
            status_code=403,
            detail="Borrado masivo de afiliados deshabilitado. Activa ALLOW_WIPE_AFILIADOS=true solo para una ventana de mantenimiento.",
        )
    if confirm != "BORRAR_AFILIADOS":
        raise HTTPException(status_code=400, detail="Confirmacion requerida para borrar todos los afiliados")
    res = await db.afiliados.delete_many({})
    return {"ok": True, "deleted": res.deleted_count}


# ---------- Eventos ----------
@api.post("/eventos", response_model=EventoOut, status_code=201)
async def crear_evento(body: EventoCreate, current=Depends(get_current_user)):
    eid = str(uuid.uuid4())
    doc = {
        "id": eid, "nombre": body.nombre, "fecha": body.fecha,
        "lugar": body.lugar or "", "descripcion": body.descripcion or "",
        "activo": True, "created_by": current["id"], "created_at": now_iso(),
    }
    await db.eventos.insert_one(doc)
    return EventoOut(**{k: v for k, v in doc.items() if k != "_id"}, total_registros=0)


@api.get("/eventos", response_model=List[EventoOut])
async def listar_eventos(current=Depends(get_current_user)):
    eventos = await db.eventos.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    out: List[EventoOut] = []
    for e in eventos:
        total = await db.registros.count_documents({"evento_id": e["id"]})
        out.append(EventoOut(
            id=e["id"], nombre=e["nombre"], fecha=e["fecha"], lugar=e.get("lugar", ""),
            descripcion=e.get("descripcion", ""), activo=e.get("activo", True),
            created_at=e["created_at"], total_registros=total,
        ))
    return out


@api.get("/eventos/{evento_id}")
async def obtener_evento(evento_id: str, current=Depends(get_current_user)):
    e = await db.eventos.find_one({"id": evento_id}, {"_id": 0})
    if not e:
        raise HTTPException(status_code=404, detail="Evento no existe")
    total = await db.registros.count_documents({"evento_id": evento_id})
    return {**e, "total_registros": total}


@api.delete("/eventos/{evento_id}")
async def eliminar_evento(evento_id: str, confirm: str = Query("", max_length=80), current=Depends(require_admin)):
    if confirm != evento_id:
        raise HTTPException(status_code=400, detail="Confirmacion requerida para eliminar el evento")
    await db.eventos.delete_one({"id": evento_id})
    r = await db.registros.delete_many({"evento_id": evento_id})
    await db.event_store.delete_many({"aggregate_id": evento_id})
    return {"ok": True, "registros_borrados": r.deleted_count}


@api.delete("/eventos/{evento_id}/registros")
async def limpiar_registros_evento(evento_id: str, confirm: str = Query("", max_length=80), current=Depends(require_admin)):
    if confirm != evento_id:
        raise HTTPException(status_code=400, detail="Confirmacion requerida para limpiar registros del evento")
    e = await db.eventos.find_one({"id": evento_id})
    if not e:
        raise HTTPException(status_code=404, detail="Evento no existe")
    r = await db.registros.delete_many({"evento_id": evento_id})
    a = await db.event_store.delete_many({"aggregate_id": evento_id})
    await db.event_store.insert_one({
        "event_type":    "EVENTO_LIMPIADO",
        "aggregate_id":  evento_id,
        "data": {
            "registros_borrados":  r.deleted_count,
            "auditoria_borrada":   a.deleted_count,
            "por": current["email"],
        },
        "timestamp":  int(datetime.now(timezone.utc).timestamp() * 1000),
        "created_at": now_iso(),
    })
    await manager.broadcast(evento_id, {
        "type":    "evento:limpiado",
        "mensaje": "Registros del evento han sido limpiados por el administrador",
        "ts":      int(datetime.now(timezone.utc).timestamp() * 1000),
    })
    return {"ok": True, "registros_borrados": r.deleted_count, "auditoria_borrada": a.deleted_count}


@api.get("/eventos/{evento_id}/estado")
async def estado_evento(evento_id: str, current=Depends(get_current_user)):
    e = await db.eventos.find_one({"id": evento_id})
    if not e:
        raise HTTPException(status_code=404, detail="Evento no existe")
    total      = await db.registros.count_documents({"evento_id": evento_id})
    today_iso  = datetime.now(timezone.utc).date().isoformat()
    hoy        = await db.registros.count_documents({"evento_id": evento_id, "created_at": {"$gte": today_iso}})
    duplicados = await db.event_store.count_documents({"aggregate_id": evento_id, "event_type": "DUPLICADO_RECHAZADO"})
    afil_cnt   = await db.registros.count_documents({"evento_id": evento_id, "es_afiliado": True})
    noaf_cnt   = await db.registros.count_documents({"evento_id": evento_id, "es_afiliado": False})
    devices    = await db.registros.distinct("device_id", {"evento_id": evento_id})
    return {
        "total":                total,
        "hoy":                  hoy,
        "duplicados_detectados": duplicados,
        "afiliados":            afil_cnt,
        "no_afiliados":         noaf_cnt,
        "dispositivos_usados":  len(devices),
        "dispositivos_activos": manager.active_count(evento_id),
    }


@api.get("/eventos/{evento_id}/registros")
async def listar_registros(evento_id: str, current=Depends(get_current_user), limit: int = 1000):
    docs = await db.registros.find({"evento_id": evento_id}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return docs


@api.get("/eventos/{evento_id}/auditoria")
async def auditoria(evento_id: str, current=Depends(get_current_user), limit: int = 300):
    docs = await db.event_store.find({"aggregate_id": evento_id}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return docs


@api.get("/eventos/{evento_id}/export")
async def export_evento(evento_id: str, current=Depends(require_admin_for_download)):
    try:
        import xlsxwriter
    except ImportError:
        raise HTTPException(status_code=500, detail="xlsxwriter no instalado")
    evento = await db.eventos.find_one({"id": evento_id}, {"_id": 0})
    if not evento:
        raise HTTPException(status_code=404, detail="Evento no existe")
    rows = await db.registros.find({"evento_id": evento_id}, {"_id": 0}).sort("created_at", 1).to_list(50000)

    buf      = io.BytesIO()
    workbook = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws       = workbook.add_worksheet("Registros")
    header_fmt  = workbook.add_format({"bold": True, "bg_color": "#009241", "color": "#FFFFFF", "border": 1, "align": "center", "valign": "vcenter"})
    cell_fmt    = workbook.add_format({"border": 1})
    cell_not_af = workbook.add_format({"border": 1, "bg_color": "#FFF5E1"})
    headers = [
        ("cedula", 14), ("primer_apellido", 18), ("segundo_apellido", 18),
        ("nombres", 22), ("nombre_completo", 28), ("genero", 8),
        ("fecha_nacimiento", 14), ("tipo_sangre", 12), ("cargo", 16),
        ("titulo", 16), ("sede", 32), ("municipio", 16), ("zona", 16),
        ("email", 24), ("celular", 14), ("es_afiliado", 12),
        ("device_id", 18), ("operador", 24), ("fecha_registro", 22),
    ]
    for c, (h, width) in enumerate(headers):
        ws.write(0, c, h, header_fmt)
        ws.set_column(c, c, width)
    ws.set_row(0, 22)
    ws.freeze_panes(1, 0)
    for r_idx, r in enumerate(rows, start=1):
        es_af = r.get("es_afiliado", False)
        fmt   = cell_fmt if es_af else cell_not_af
        ws.write(r_idx,  0, r.get("cedula", ""),          fmt)
        ws.write(r_idx,  1, r.get("primer_apellido", ""), fmt)
        ws.write(r_idx,  2, r.get("segundo_apellido", ""),fmt)
        ws.write(r_idx,  3, r.get("nombres", ""),         fmt)
        ws.write(r_idx,  4, r.get("nombre", ""),          fmt)
        ws.write(r_idx,  5, r.get("genero", ""),          fmt)
        ws.write(r_idx,  6, r.get("fecha_nacimiento", ""),fmt)
        ws.write(r_idx,  7, r.get("tipo_sangre", ""),     fmt)
        ws.write(r_idx,  8, r.get("cargo", ""),           fmt)
        ws.write(r_idx,  9, r.get("titulo", ""),          fmt)
        ws.write(r_idx, 10, r.get("sede", ""),            fmt)
        ws.write(r_idx, 11, r.get("municipio", ""),       fmt)
        ws.write(r_idx, 12, r.get("zona", ""),            fmt)
        ws.write(r_idx, 13, r.get("email", ""),           fmt)
        ws.write(r_idx, 14, r.get("celular", ""),         fmt)
        ws.write(r_idx, 15, "SI" if es_af else "NO",      fmt)
        ws.write(r_idx, 16, r.get("device_id", ""),       fmt)
        ws.write(r_idx, 17, r.get("operator_email", ""),  fmt)
        ws.write(r_idx, 18, r.get("created_at", ""),      fmt)
    ws.autofilter(0, 0, max(1, len(rows)), len(headers) - 1)

    meta = workbook.add_worksheet("Resumen")
    meta.set_column(0, 0, 22)
    meta.set_column(1, 1, 44)
    meta.write(0, 0, "Campo", header_fmt)
    meta.write(0, 1, "Valor", header_fmt)
    afiliados_count = sum(1 for r in rows if r.get("es_afiliado"))
    no_af_count     = len(rows) - afiliados_count
    meta_rows = [
        ("Evento",          evento.get("nombre", "")),
        ("Fecha evento",    evento.get("fecha", "")),
        ("Lugar",           evento.get("lugar", "")),
        ("Descripción",     evento.get("descripcion", "")),
        ("Total registros", len(rows)),
        ("Afiliados",       afiliados_count),
        ("No afiliados",    no_af_count),
        ("Generado",        datetime.now(timezone.utc).isoformat()),
        ("Generado por",    current.get("email", "")),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        meta.write(i, 0, k,      cell_fmt)
        meta.write(i, 1, str(v), cell_fmt)

    workbook.close()
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() else "_" for c in evento["nombre"])[:40]
    fname     = f"registros_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.patch("/eventos/{evento_id}/registros/{registro_id}")
async def admin_update_registro(evento_id: str, registro_id: str, body: RegistroUpdate, current=Depends(require_admin)):
    data = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not data:
        raise HTTPException(status_code=400, detail="Sin cambios")
    res = await db.registros.update_one({"id": registro_id, "evento_id": evento_id}, {"$set": data})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return {"ok": True}


@api.delete("/eventos/{evento_id}/registros/{registro_id}")
async def admin_delete_registro(evento_id: str, registro_id: str, current=Depends(require_admin)):
    reg = await db.registros.find_one({"id": registro_id, "evento_id": evento_id})
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    await db.registros.delete_one({"id": registro_id})
    await db.event_store.insert_one({
        "event_type":   "REGISTRO_ELIMINADO",
        "aggregate_id": evento_id,
        "data": {"registro_id": registro_id, "cedula": reg.get("cedula"), "por": current["email"]},
        "timestamp":  int(datetime.now(timezone.utc).timestamp() * 1000),
        "created_at": now_iso(),
    })
    await manager.broadcast(evento_id, {
        "type":        "registro:eliminado",
        "registro_id": registro_id,
        "cedula":      reg.get("cedula"),
        "ts":          int(datetime.now(timezone.utc).timestamp() * 1000),
    })
    return {"ok": True}


# ---------- Escaneo ----------
@api.post("/registros/escanear")
async def escanear(
    body: EscanearRequest,
    current=Depends(get_current_user),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
):
    parsed = parse_pdf417(body.raw_barcode) if body.raw_barcode else {}
    client_cedula = clean_cedula(body.cedula)
    barcode_cedula = trusted_barcode_cedula(parsed)
    cedula_norm = barcode_cedula or client_cedula
    if not cedula_norm:
        raise HTTPException(status_code=400, detail="Cédula inválida")
    if barcode_cedula and client_cedula and barcode_cedula != client_cedula:
        logger.info(
            "Cedula corregida desde raw_barcode: cliente=%s servidor=%s formato=%s",
            client_cedula,
            barcode_cedula,
            parsed.get("formato_detectado"),
        )

    evento = await db.eventos.find_one({"id": body.evento_id})
    if not evento:
        raise HTTPException(status_code=404, detail="Evento no existe")
    if not evento.get("activo", True):
        raise HTTPException(status_code=400, detail="Evento inactivo")

    dedup_hash = compute_dedup_hash(cedula_norm, body.evento_id, body.device_id, body.nonce)
    idem_key   = idempotency_key or dedup_hash

    cached = await db.deduplication.find_one({"idempotency_key": idem_key}, {"_id": 0})
    if cached:
        return cached["response"]

    existing = await db.registros.find_one({"evento_id": body.evento_id, "cedula": cedula_norm}, {"_id": 0})
    if existing:
        await db.event_store.insert_one({
            "event_type":   "DUPLICADO_RECHAZADO",
            "aggregate_id": body.evento_id,
            "data": {
                "cedula":               cedula_norm,
                "device_id":            body.device_id,
                "operator_email":       current["email"],
                "registro_original_id": existing["id"],
            },
            "timestamp":  int(datetime.now(timezone.utc).timestamp() * 1000),
            "created_at": now_iso(),
            "dedup_hash": dedup_hash,
        })
        await manager.broadcast(body.evento_id, {
            "type":             "registro:duplicado",
            "cedula":           cedula_norm,
            "device_id":        body.device_id,
            "registro_original": existing,
            "mensaje":          f"Cédula {cedula_norm} ya registrada",
            "ts":               int(datetime.now(timezone.utc).timestamp() * 1000),
        })
        response = {
            "ok": False,
            "error": "DUPLICADO",
            "mensaje": "Cédula ya registrada en este evento",
            "registro_existente": existing,
        }
        await db.deduplication.update_one(
            {"idempotency_key": idem_key},
            {"$set": {"idempotency_key": idem_key, "response": response, "created_at": now_iso()}},
            upsert=True,
        )
        raise HTTPException(status_code=409, detail=response)

    afiliado   = await db.afiliados.find_one({"cedula": cedula_norm}, {"_id": 0})
    es_afiliado = bool(afiliado)

    def pref(client_val: str, afi_key: str, parsed_key: str = "") -> str:
        if client_val:
            return client_val.strip()
        if afiliado and afi_key:
            v = (afiliado.get(afi_key) or "").strip()
            if v:
                return v
        if parsed_key:
            return (parsed.get(parsed_key) or "").strip()
        return ""

    composed_nombre = ""
    if parsed.get("primer_apellido") or parsed.get("nombres"):
        composed_nombre = " ".join(filter(None, [
            parsed.get("nombres", ""),
            parsed.get("primer_apellido", ""),
            parsed.get("segundo_apellido", ""),
        ])).strip()
    composed_nombre = clean_display_text(composed_nombre)

    registro_id = str(uuid.uuid4())
    ts_ms       = int(datetime.now(timezone.utc).timestamp() * 1000)
    doc = {
        "id":               registro_id,
        "evento_id":        body.evento_id,
        "cedula":           cedula_norm,
        "nombre":           clean_display_text(pref(body.nombre or "", "nombre") or composed_nombre),
        "primer_apellido":  (body.primer_apellido  or parsed.get("primer_apellido")  or "").strip(),
        "segundo_apellido": (body.segundo_apellido or parsed.get("segundo_apellido") or "").strip(),
        "nombres":          (body.nombres          or parsed.get("nombres")          or "").strip(),
        "genero":           (body.genero           or parsed.get("genero")           or "").strip(),
        "fecha_nacimiento": (
            body.fecha_nacimiento
            or parsed.get("fecha_nacimiento")
            or (afiliado.get("fecha_nac") if afiliado else "")
            or ""
        ).strip(),
        "fecha_expiracion": (body.fecha_expiracion or parsed.get("fecha_expiracion") or "").strip(),
        "tipo_sangre":  (body.tipo_sangre  or parsed.get("tipo_sangre") or "").strip(),
        "sede":         pref(body.sede     or "", "sede"),
        "municipio":    pref(body.municipio or "", "municipio"),
        "cargo":        pref(body.cargo    or "", "cargo"),
        "zona":         (afiliado.get("zona")   if afiliado else "") or "",
        "titulo":       (afiliado.get("titulo") if afiliado else "") or "",
        "email":        (afiliado.get("email")  if afiliado else "") or "",
        "celular":      (afiliado.get("celular") if afiliado else "") or "",
        "es_afiliado":  es_afiliado,
        "device_id":    body.device_id,
        "operator_email": current["email"],
        "operator_id":    current["id"],
        "timestamp":    body.timestamp or ts_ms,
        "nonce":        body.nonce,
        "dedup_hash":   dedup_hash,
        "created_at":   now_iso(),
    }
    try:
        await db.registros.insert_one(dict(doc))
    except DuplicateKeyError:
        winner = await db.registros.find_one({"evento_id": body.evento_id, "cedula": cedula_norm}, {"_id": 0})
        await db.event_store.insert_one({
            "event_type":   "DUPLICADO_RECHAZADO",
            "aggregate_id": body.evento_id,
            "data": {"cedula": cedula_norm, "device_id": body.device_id, "operator_email": current["email"], "race": True},
            "timestamp": ts_ms, "created_at": now_iso(), "dedup_hash": dedup_hash,
        })
        await manager.broadcast(body.evento_id, {
            "type":             "registro:duplicado",
            "cedula":           cedula_norm,
            "device_id":        body.device_id,
            "registro_original": winner,
            "mensaje":          f"Cédula {cedula_norm} ya registrada (race)",
            "ts":               ts_ms,
        })
        response = {"ok": False, "error": "DUPLICADO_RACE", "registro_existente": winner}
        await db.deduplication.update_one(
            {"idempotency_key": idem_key},
            {"$set": {"idempotency_key": idem_key, "response": response, "created_at": now_iso()}},
            upsert=True,
        )
        raise HTTPException(status_code=409, detail=response)

    payload_out = {k: v for k, v in doc.items() if k not in ("_id", "nonce", "dedup_hash", "operator_id")}
    await db.event_store.insert_one({
        "event_type":   "REGISTRO_CREADO",
        "aggregate_id": body.evento_id,
        "data":         payload_out,
        "timestamp":    ts_ms,
        "created_at":   now_iso(),
        "dedup_hash":   dedup_hash,
    })
    await manager.broadcast(body.evento_id, {"type": "registro:nuevo", "registro": payload_out, "ts": ts_ms})

    response = {"ok": True, "registro": payload_out, "afiliado_encontrado": es_afiliado}
    await db.deduplication.update_one(
        {"idempotency_key": idem_key},
        {"$set": {"idempotency_key": idem_key, "response": response, "created_at": now_iso()}},
        upsert=True,
    )
    return response


# ---------- OCR ----------
class OcrRequest(BaseModel):
    image_base64: str = Field(min_length=10)
    force_gemini: bool = False
    prefer_mrz: bool = False


@api.post("/ocr/cedula")
async def api_ocr_cedula(body: OcrRequest, current=Depends(get_current_user)):
    result   = await ocr_from_base64_async(body.image_base64, force_gemini=body.force_gemini)
    parsed   = result.get("parsed") or {}
    afiliado = None
    if result.get("cedula"):
        afiliado = await db.afiliados.find_one({"cedula": result["cedula"]}, {"_id": 0})
    return {
        "ok":      bool(result.get("cedula")),
        "cedula":  result.get("cedula") or "",
        "texto":   (result.get("texto_completo") or "")[:300],
        "texto_completo": result.get("texto_completo") or "",
        "pipeline_usado": result.get("pipeline_usado") or "",
        "parsed":  parsed,
        "raw_mrz": parsed.get("raw_mrz") or [],
        "mrz_valido": bool(parsed.get("mrz_valido")),
        "gemini": result.get("gemini") or {},
        "afiliado": afiliado,
        "error":   result.get("error", ""),
    }


@api.post("/barcode/cedula")
async def api_barcode_cedula(body: OcrRequest, current=Depends(get_current_user)):
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(None, decode_identity_document_from_base64, body.image_base64, body.prefer_mrz)
    afiliado = None
    if result.get("cedula"):
        afiliado = await db.afiliados.find_one({"cedula": result["cedula"]}, {"_id": 0})
    return {
        "ok": bool(result.get("cedula")),
        "cedula": result.get("cedula") or "",
        "raw": result.get("raw") or "",
        "parsed": result.get("parsed") or {},
        "format": result.get("format") or "",
        "source": result.get("source") or "",
        "candidates": result.get("candidates") or [],
        "raw_mrz": result.get("raw_mrz") or (result.get("parsed") or {}).get("raw_mrz") or [],
        "mrz_valido": bool(result.get("mrz_valido") or (result.get("parsed") or {}).get("mrz_valido")),
        "afiliado": afiliado,
        "error": result.get("error", ""),
    }


# ---------- Health ----------
@api.get("/")
async def root():
    return {"service": "CedulaScan Pro", "status": "ok", "version": "2.1.0"}


@api.get("/health/dependencies")
async def health_dependencies():
    return {
        "ok": True,
        "ocr": ocr_dependency_status(),
        "barcode": barcode_dependency_status(),
        "gemini": gemini_dependency_status(),
    }

# ======================================================================
# 🚀 ENDPOINT CRÍTICO: PROCESAMIENTO HÍBRIDO (TIEMPO REAL + RESPALDO)
# ======================================================================
@api.post("/ocr/cedula/hybrid-register")
async def ocr_cedula(payload: dict, current=Depends(get_current_user)):
    """
    Procesa escaneos de cédula. Soporta ráfaga rápida por texto plano (is_raw_string)
    o procesamiento de respaldo mediante fotos en Base64. Garantiza deduplicación en Mongo.
    """
    evento_id = payload.get("evento_id")
    device_id = payload.get("device_id", "DESKTOP_DEV")
    nonce = payload.get("nonce", str(uuid.uuid4())[:8])
    
    if not evento_id:
        raise HTTPException(status_code=400, detail="El campo 'evento_id' es obligatorio")

    # Verificar que el evento esté activo
    evento = await db.eventos.find_one({"id": evento_id})
    if not evento:
        raise HTTPException(status_code=404, detail="El evento especificado no existe")

    raw_text = ""
    metodo = "imagen_base64"

    # ------------------------------------------------------------------
    # MODO 1: ⚡ TIEMPO REAL - EXPO YA LEYÓ EL TEXTO BINARIO CRUDO
    # ------------------------------------------------------------------
    if payload.get("is_raw_string"):
        raw_text = payload.get("raw_data", "")
        metodo = "texto_directo"
        print(f"\n[BACKEND-⚡] ¡Recibido texto directo desde el lector nativo!")
        
        if not raw_text:
            raise HTTPException(status_code=400, detail="El texto de barras enviado está vacío")
            
        try:
            parsed = parse_pdf417(raw_text)
            cedula = trusted_barcode_cedula(parsed) or parsed.get("cedula") or extract_cedula(raw_text) or ""
        except Exception as pe:
            print(f"[BACKEND-❌] Error parseando texto de barras: {pe}")
            cedula = ""
            parsed = {}

    # ------------------------------------------------------------------
    # MODO 2: 📸 RESPALDO - PROCESAR IMAGEN BASE64
    # ------------------------------------------------------------------
    else:
        image_base64 = payload.get("image", "") or payload.get("image_base64", "")
        if not image_base64:
            raise HTTPException(status_code=400, detail="No se proporcionó texto de barras ni imagen Base64")
            
        print("\n[BACKEND-📸] Procesando por respaldo de imagen...")
        resultado_img = decode_identity_document_from_base64(image_base64)

        if not resultado_img.get("ok"):
            ocr_result = await ocr_from_base64_async(image_base64, force_gemini=False)
            if ocr_result.get("cedula"):
                resultado_img = {
                    "ok": True,
                    "cedula": ocr_result.get("cedula", ""),
                    "raw": ocr_result.get("texto_completo", ""),
                    "parsed": ocr_result.get("parsed", {}) or {},
                    "source": "ocr_front",
                    "format": "ocr",
                    "raw_mrz": ocr_result.get("raw_mrz", []) or [],
                    "mrz_valido": bool(ocr_result.get("mrz_valido")),
                    "error": "",
                }
            else:
                return {
                    "ok": False,
                    "error": resultado_img.get("error") or ocr_result.get("error") or "No se detectó el código PDF417",
                }

        cedula = resultado_img.get("cedula")
        raw_text = resultado_img.get("raw", "")
        parsed = resultado_img.get("parsed", {})

    # ------------------------------------------------------------------
    # LÓGICA DE NEGOCIO: VALIDACIÓN, DEDUPLICACIÓN Y GUARDADO EN MONGO
    # ------------------------------------------------------------------
    cedula_limpia = trusted_barcode_cedula(parsed) or clean_cedula(cedula)
    if not cedula_limpia:
        return {"ok": False, "error": "Código decodificado pero el número de cédula no es válido para Colombia"}

    dedup_hash = compute_dedup_hash(cedula_limpia, evento_id, device_id, nonce)

    # Verificar duplicados
    ya_existe = await db.registros.find_one({"evento_id": evento_id, "cedula": cedula_limpia})
    if ya_existe:
        print(f"[BACKEND-⚠️] Cédula {cedula_limpia} rechazada por duplicado en el evento.")
        
        await db.event_store.insert_one({
            "event_type": "DUPLICADO_RECHAZADO",
            "aggregate_id": evento_id,
            "data": {"cedula": cedula_limpia, "device_id": device_id, "operador": current["email"]},
            "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000),
            "created_at": now_iso()
        })
        
        await manager.broadcast(evento_id, {
            "type": "registro:duplicado",
            "cedula": cedula_limpia,
            "mensaje": f"La cédula {cedula_limpia} ya fue registrada",
            "ts": int(datetime.now(timezone.utc).timestamp() * 1000)
        })
        
        return {"ok": False, "error": "Esta cédula ya fue registrada en este evento", "duplicado": True, "cedula": cedula_limpia}

    # Cruzar datos con Afiliados
    afiliado_db = await db.afiliados.find_one({"cedula": cedula_limpia}, {"_id": 0})
    es_afiliado = isinstance(afiliado_db, dict)

    registro_id = str(uuid.uuid4())
    doc_registro = {
        "id": registro_id,
        "evento_id": evento_id,
        "cedula": cedula_limpia,
        "es_afiliado": es_afiliado,
        "device_id": device_id,
        "operator_id": current["id"],
        "operator_email": current["email"],
        "metodo_captura": metodo,
        "dedup_hash": dedup_hash,
        "created_at": now_iso(),
        
        "primer_apellido": clean_display_text(parsed.get("primer_apellido", "")),
        "segundo_apellido": clean_display_text(parsed.get("segundo_apellido", "")),
        "nombres": clean_display_text(parsed.get("nombres", "")),
        "genero": parsed.get("genero", ""),
        "fecha_nacimiento": parsed.get("fecha_nacimiento", ""),
        "tipo_sangre": parsed.get("tipo_sangre", ""),
        
        "nombre": afiliado_db.get("nombre") if es_afiliado else f"{parsed.get('nombres', '')} {parsed.get('primer_apellido', '')}".strip(),
        "sede": afiliado_db.get("sede", "") if es_afiliado else payload.get("sede", ""),
        "municipio": afiliado_db.get("municipio", "") if es_afiliado else payload.get("municipio", ""),
        "zona": afiliado_db.get("zona", "") if es_afiliado else "",
        "cargo": afiliado_db.get("cargo", "") if es_afiliado else payload.get("cargo", ""),
        "titulo": afiliado_db.get("titulo", "") if es_afiliado else "",
        "email": afiliado_db.get("email", "") if es_afiliado else "",
        "celular": afiliado_db.get("celular", "") if es_afiliado else ""
    }

    try:
        await db.registros.insert_one(doc_registro)
        
        await db.event_store.insert_one({
            "event_type": "AFILIADO_REGISTRADO" if es_afiliado else "INVITADO_REGISTRADO",
            "aggregate_id": evento_id,
            "data": {"registro_id": registro_id, "cedula": cedula_limpia, "es_afiliado": es_afiliado},
            "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000),
            "created_at": now_iso()
        })
        
        if "_id" in doc_registro:
            del doc_registro["_id"]
            
        await manager.broadcast(evento_id, {
            "type": "registro:nuevo",
            "data": doc_registro,
            "ts": int(datetime.now(timezone.utc).timestamp() * 1000)
        })
        
        print(f"[BACKEND-✅] Registro exitoso en MongoDB. Cédula: {cedula_limpia}")
        return {"ok": True, "cedula": cedula_limpia, "es_afiliado": es_afiliado, "registro": doc_registro}

    except DuplicateKeyError:
        return {"ok": False, "error": "Petición duplicada en proceso", "duplicado": True}


# ---------- Incluir Rutas del API Router ----------
app.include_router(api)


# ---------- Sistema Unificado de WebSockets ----------
@app.websocket("/api/eventos/{evento_id}")
async def ws_evento(websocket: WebSocket, evento_id: str, token: str = "", device_id: str = ""):
    """
    Maneja la suscripción en tiempo real de los clientes al evento utilizando la lógica original.
    """
    try:
        payload = decode_token(token)
        user_email = payload.get("email", "anon")
    except Exception:
        await websocket.close(code=4401)
        return

    evento = await db.eventos.find_one({"id": evento_id})
    if not evento:
        await websocket.close(code=4404)
        return

    dev = device_id or f"dev_{uuid.uuid4().hex[:8]}"
    await manager.connect(websocket, evento_id, dev)
    
    total = await db.registros.count_documents({"evento_id": evento_id})
    
    await websocket.send_json({
        "type": "estado:inicial",
        "total": total,
        "activos": manager.active_count(evento_id),
        "operator": user_email,
        "ts": int(datetime.now(timezone.utc).timestamp() * 1000),
    })
    
    try:
        while True:
            msg = await websocket.receive_json()
            if msg.get("type") == "ping":
                await websocket.send_json({
                    "type": "pong",
                    "ts": int(datetime.now(timezone.utc).timestamp() * 1000),
                })
    except WebSocketDisconnect:
        await manager.disconnect(websocket)
    except Exception as e:
        logger.warning("WS error: %s", e)
        await manager.disconnect(websocket)


if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.environ.get("PORT", 8001))
    uvicorn.run("server:app", host="0.0.0.0", port=port, reload=False)
